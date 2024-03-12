import os.path

import openpyxl as xl
from lxml import etree as et
from lxml.etree import Element, ElementTree

from libs.Common.commonVar import common


class langsDB(common):
    def __init__(self):
        super().__init__()

    def runExtractlang(self):
        print("runExtractlang 시작")

        # 1. 로드할 엑셀 경로 생성
        self.langF = self.resourceDir / "excel-template/language.xlsx"
        print("self.langF:", self.langF)

        # 2. 엑셀 파일 읽기
        wb = xl.load_workbook(self.langF)
        # 3. 엑셀 내 모든 시틈 이름 추출
        self.sheetnames = wb.sheetnames
        # 4. 시트 개수 파악
        self.sheetCnt = len(self.sheetnames)

        for i in range(self.sheetCnt):
            # 5. 시트 로드
            self.cursheet = wb.worksheets[i]
            # 6. 현재 시트 이름 추출
            self.sheetname = self.cursheet.title.replace(" ", "_").lower()
            # print(f'{self.sheetname=}')
            # 7. excel 데이터를 sheet 별 xml 파일로 저장하기 위해 root 태그 생성
            self.root = Element("root")
            self.root.set("filename", self.sheetname)
            # 8. rows : 시트내 모든 Cell을 행별로 추출
            self.rowrng = self.cursheet.rows
            # 시작 row, column 설정
            self.startRow = 1
            self.startCol = 1

            self.headCnt = 0
            for row in enumerate(self.rowrng):
                # 9. 인덱스 추출
                self.rowIdx = row[0]
                self.curRow = row[1]

                self.listitem = Element("listitem")

                for cellinfo in self.curRow:
                    self.colIdx = cellinfo.column
                    self.cellval = cellinfo.value

                    if self.colIdx > 1:
                        if self.rowIdx == 1:
                            # head 행에서 비어 있지 않은 Cell 개수 파악 하기
                            if any(cell.value is not None for cell in self.curRow):
                                self.headCnt += 1

                        elif self.rowIdx > 1:
                            self.root.append(self.listitem)
                            if self.sheetname == "language":
                                if self.colIdx == 2:
                                    self.listitem.set(
                                        "language", self.cellval.replace(" ", "_")
                                    )

                                elif self.colIdx == 3:
                                    self.listitem.set(
                                        "ISOCode", self.cellval.replace(" ", "_")
                                    )

                            elif self.sheetname == "type":
                                self.listitem.set(
                                    "type", self.cellval.replace(" ", "_")
                                )

            # root 태그를 DOM 트리 구조로 변환
            self.doc = ElementTree(self.root)

            # 저장될 폴더 생성
            self.saveExcelF = os.path.join(self.resourceDir, self.sheetname + ".xml")
            self.doc.write(
                self.saveExcelF,
                pretty_print=True,
                encoding="utf8",
                method="xml",
                xml_declaration=True,
            )

    def loadType(self):
        print("loadType 시작")

        # type 속성값들을 수집할 리스트 객체 생성
        self.typeL = []

        # 1. type.xml파일 경로 선언
        typeF = os.path.join(common.resourceDir, "type.xml")
        # 2. xml 파서 방법 지정
        parser = et.XMLParser(ns_clean=True, encoding="utf8", recover=True)
        # 3. xml 문서 파서
        self.doc = et.parse(typeF, parser)
        # 4. xml 문서내 root 태그에 접근
        self.root = self.doc.getroot()
        # 5. root 태그 하위 자식 태그 접근
        for child in self.root.iter("listitem"):
            attrType = child.get("type")
            # 6. 리스트 객체에 할당
            self.typeL.append(attrType)

    def loadLangs(self):
        print("loadLangs 시작")

        # type 속성값들을 수집할 리스트 객체 생성
        self.langMap = {}

        # 1. type.xml파일 경로 선언
        typeF = os.path.join(common.resourceDir, "language.xml")
        # 2. xml 파서 방법 지정
        parser = et.XMLParser(ns_clean=True, encoding="utf8", recover=True)
        # 3. xml 문서 파서
        self.doc = et.parse(typeF, parser)
        # 4. xml 문서내 root 태그에 접근
        self.root = self.doc.getroot()
        # 5. root 태그 하위 자식 태그 접근
        for child in self.root.iter("listitem"):
            attrLang = child.get("language")
            attrISO = child.get("ISOCode")
            # 6. 딕셔너리 객체에 할당
            self.langMap[attrLang] = attrISO

    def getListNmap(self):
        return self.typeL, self.langMap
