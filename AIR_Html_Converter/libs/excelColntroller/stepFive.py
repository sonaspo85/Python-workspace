# -*- coding: utf-8 -*-
import os
import re
import traceback

import openpyxl as xl
from lxml.etree import Element, SubElement, ElementTree

from libs.Common.commonVar import common


class stepFive(common):
    def __init__(self, pbar1, lb4):
        super().__init__()
        self.pbar1 = pbar1
        self.lb4 = lb4

    def runStepFive(self):
        print("runStepFive 시작")

        try:
            # 1. 경로내 모든 파일 추출
            self.files = os.listdir(self.excelTemplsDir)
            # print('self.files:', self.files)
            # 2. 반복
            for file in self.files:
                self.filename = file
                self.abspath = os.path.abspath(
                    os.path.join(self.excelTemplsDir, self.filename)
                )
                # print('self.abspath:', self.abspath)
                # 3. 엑셀 파일인 경우
                if (
                    self.filename.endswith(".xlsx")
                    and self.filename.find("template1") > -1
                ):
                    # 4. 엑셀 파일 로드
                    self.wb = xl.load_workbook(self.abspath)
                    # 5. 엑셀 내 모든 시트 이름 추출
                    self.sheetnames = self.wb.sheetnames
                    # 엑셀 시트 개수
                    self.sheetcnt = len(self.sheetnames)

                    for i in range(self.sheetcnt):
                        # 6. 시트 로드
                        self.cursheet = self.wb.worksheets[i]
                        # 현재 시트 이름 추출
                        self.cursheetname = self.cursheet.title.replace(
                            " ", "_"
                        ).lower()
                        print(f"{self.cursheetname=}")

                        # 7. sheet를 xml로 추출 하기 위해 root 태그 생성
                        self.root = Element("root")
                        self.root.set("filename", self.cursheetname)

                        self.startRow = 1
                        self.startCol = 1

                        # 8. rows : 시트내 모든 cell을 행별로 추출
                        self.rowrng = self.cursheet.rows

                        # head Cell값을 저장할 list 객체
                        self.headCellL = []

                        for row in enumerate(self.rowrng):
                            # 9. 현재 row 위치 인덱스 추출
                            self.rowIdx = row[0]
                            self.curRow = row[1]

                            self.cnt1 = 0
                            # 10. head cell 값 추출 하여 self.headCellL 리스트 객체에 보관 (63-76)
                            for cellinfo in self.curRow:
                                self.colIdx = cellinfo.column  # column 위치
                                self.cellval = cellinfo.value  # 현재 cell 값
                                # print(f'{self.rowIdx} : {self.colIdx} = {self.cellval}')

                                # 11. 실제 시작 row, column 위치 부터 추출
                                if (
                                    self.rowIdx > self.startRow
                                    and self.colIdx > self.startCol
                                ):
                                    if self.rowIdx == 2:
                                        self.headCell = re.sub(
                                            "[()]", "_", self.cellval
                                        )
                                        self.headCell2 = re.sub("_$", "", self.headCell)
                                        # 12. head cell값을 list 객체로 수집
                                        self.headCellL.append(self.headCell2)

                            # 13. body row 인 경우
                            if self.rowIdx > self.startRow + 1:
                                # 14. 한번씩 돌때 마다 태그 생성
                                self.listitem = Element("listitem")
                                self.root.append(self.listitem)

                                # body row 인 경우 반복
                                for cellinfo in self.curRow:
                                    self.colIdx = cellinfo.column  # column 위치
                                    self.cellval = cellinfo.value  # 현재 cell 값
                                    self.attrCls = f"{self.rowIdx+1}:{self.colIdx}"

                                    if self.colIdx > self.startCol:
                                        self.item = SubElement(
                                            self.listitem,
                                            self.headCellL[self.cnt1],
                                            {"class": self.attrCls},
                                        ).text = self.cellval
                                        self.cnt1 += 1

                        # root 태그를 DOM 트리 구조로 변환
                        self.doc = ElementTree(self.root)

                        # 저장될 폴더 생성
                        self.saveDir = os.path.join(self.tempDir, "excelTempls")
                        if not os.path.exists(self.saveDir):
                            os.makedirs(self.saveDir)

                        # xml 파일로 저장
                        self.saveExcelP = os.path.join(
                            self.saveDir, self.cursheetname + ".xml"
                        )
                        self.doc.write(
                            self.saveExcelP,
                            pretty_print=True,
                            encoding="utf8",
                            method="xml",
                        )

                        self.pbar1.setValue(50)
                        self.lb4.setText("template1.xlsx xml 변환 완료")

        except Exception as e:
            msg = "resource/excel-template 내 엑셀 파일 읽기 실패"

            raise Exception(msg)
