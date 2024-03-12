# -*- coding: utf-8 -*-
import os
import re

import openpyxl as xl
from lxml.etree import Element, ElementTree, SubElement

from libs.Common.commonVar import common


class stepFour(common):
    def __init__(self):
        super().__init__()

    def runStepFour(self):
        print("runStepFour 시작")
        # 1. excel-template 경로 생성
        # self.excelTemplsDir = os.path.join(self.resourceDir, 'excel-template')
        self.excelTemplsDir = os.path.join(self.excelTemplsDir)
        print("self.excelTemplsDir: ", self.excelTemplsDir)

        # 2. excel-template 하위 디렉토리 탐색
        self.excelDirs = os.listdir(self.excelTemplsDir)

        for excel in self.excelDirs:
            # 3. 절대 경로 추출
            self.abpath = os.path.join(self.excelTemplsDir, excel)

            if os.path.isfile(self.abpath) and excel.endswith(".xlsx"):
                # 4. Excel 파일 로드
                wb = xl.load_workbook(self.abpath)
                # 5. 엑셀내 모든 시트 이름 추출
                self.sheetnameL = wb.sheetnames
                # 6. 시트 개수
                self.sheetCnt = len(self.sheetnameL)
                # 7. 시트 반복
                for i in range(self.sheetCnt):
                    # 8. 시트 로드
                    self.curSheet = wb.worksheets[i]
                    # 9. 현재 시트 이름 추출
                    self.sheetName = self.curSheet.title.replace(" ", "_").lower()

                    # 10. excel 데이터를 sheet 별로 xml 파일로 저장하기 위해 root 태그 생성
                    self.root = Element("root")
                    self.root.set("filename", self.sheetName)

                    # 11. rows: 시트내 모든 Cell을 행별로 추출
                    self.rowrng = self.curSheet.rows

                    # 12. 시작 row 및 시작 colunm 설정
                    self.startRow = 2
                    self.startCol = 1
                    # head Cell값을 저장할 list 객체
                    self.headCellL = []

                    # 13. row별로 수집된 self.rowrng 객체를 반복
                    for row in enumerate(self.rowrng):
                        # 14. row 인덱스 추출
                        self.rowIdx = row[0]
                        self.curRow = row[1]
                        # 15. head 행인 경우
                        if self.rowIdx == self.startRow:
                            # 16. head 행의 각 cell 반복
                            for cellinfo in self.curRow:
                                self.colIdx = cellinfo.column
                                self.cellval = cellinfo.value

                                # 17. column 시작 위치 체크
                                if self.colIdx > self.startCol:
                                    # re.sub(): 정규식 패턴과 일치하는 문자열을 다른 문자열로 교체
                                    self.headcell = re.sub("[()]", "_", self.cellval)
                                    # print(self.headcell)
                                    # 18. head cell값을 list 객체로 수집
                                    self.headCellL.append(self.headcell)

                        # 19. body 행인 경우
                        elif self.startRow < self.rowIdx:
                            # 20. head Cell의 개수 파악
                            self.totalCellCnt = len(self.headCellL)

                            # 21. 행이 한번씩 돌때마다 태그 생성
                            self.listitem = Element("listitem")
                            self.root.append(self.listitem)

                            for pos, item in enumerate(self.headCellL):
                                # 23. excel행과 열을 속성으로 생성
                                self.attrCls = f"{self.rowIdx + 1}:{pos + 1}"
                                # 24. sheet.cell(row, column) 함수로 cell값 읽기
                                self.curCell = self.curSheet.cell(
                                    self.rowIdx + 1, pos + 2
                                ).value

                                # 현재 Cell이 위치한 Cell 객체 생성
                                self.cellObj = self.curSheet.cell(
                                    self.rowIdx + 1, pos + 2
                                )
                                # 생성한 Cell 객체로 부터 현재 cell 번호 추출
                                self.colpos = self.cellObj.column
                                self.rowpos = self.cellObj.row

                                # 현재 Cell이 위치한 좌표 추출
                                self.curCoordinate = self.cellObj.coordinate

                                if self.curCell == None:
                                    # merged__cells: 병합된 셀 객체를 반환
                                    for crange in self.curSheet.merged_cells:
                                        # bounds : 병합된 Cell의 4개 모서리를 튜퓨 타입으로 반환 한다.
                                        colow, rowlow, colhigh, rowhigh = crange.bounds

                                        # 첫번째 Cell 위치의 값 추출
                                        top_value = self.curSheet.cell(
                                            rowlow, colow
                                        ).value
                                        # print(f'{top_value=}')
                                        # 병합된 Cell이 None 인 경우 병합된 첫번째 Cell값 할당
                                        if (
                                            rowlow <= self.rowpos
                                            and self.rowpos <= rowhigh
                                            and colow <= self.colpos
                                            and self.colpos <= colhigh
                                        ):
                                            self.curCell = top_value
                                            print(
                                                f"[{self.rowpos}:{self.colpos}]",
                                                "병합된 Cell:",
                                                self.curCell,
                                            )

                                            break

                                    if self.curCell == None:
                                        self.curCell = ""

                                SubElement(
                                    self.listitem, item, {"class": self.attrCls}
                                ).text = self.curCell

                    # 26. root 태그를 DOM 트리 구조로 변환
                    self.doc = ElementTree(self.root)

                    # 27. 저장될 폴더 생성
                    if not os.path.exists(os.path.join(self.tempDir, "excelTempls")):
                        os.makedirs(os.path.join(self.tempDir, "excelTempls"))

                    # 28. xml 파일로 저장
                    self.saveExcelP = os.path.join(
                        self.tempDir, "excelTempls", self.sheetName + ".xml"
                    )
                    self.doc.write(
                        self.saveExcelP,
                        pretty_print=True,
                        encoding="utf8",
                        method="xml",
                    )
